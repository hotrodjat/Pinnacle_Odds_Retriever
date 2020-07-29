from bs4 import BeautifulSoup
from selenium import webdriver
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from dateutil.tz import tzutc
import time as tm
import datetime
import dateutil.parser
import re

times_list = []
temp_driver = None


def update_soup():
    driver = webdriver.Chrome(executable_path=".\chromedriver.exe")
    driver.maximize_window()
    driver.get("https://www.pinnacle.com/en/soccer/italy-serie-a/matchups")
    tm.sleep(15.0)
    html = driver.page_source
    driver.quit()
    del driver
    return BeautifulSoup(html, 'lxml')


def start_temp_driver():
    global temp_driver
    temp_driver = webdriver.Chrome(executable_path=".\chromedriver.exe")
    temp_driver.maximize_window()


def close_temp_driver():
    global temp_driver
    temp_driver.quit()
    del temp_driver


def pull_event_odds(event_url):
    global temp_driver
    temp_driver.get(event_url)
    tm.sleep(13.0)
    html = temp_driver.page_source
    return BeautifulSoup(html, 'lxml')


def format_text(ws):
    ws.cell(row=ws.max_row, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=3)


def print_odds(moneyline, handicap, total, event_time, time_until):
    wb = load_workbook('ODDS_OUTPUT.xlsx')
    ws = wb['Sheet1']

    ws.cell(row=ws.max_row + 3, column=1).value = f"{moneyline[0][0]} Vs. {moneyline[0][2]}"
    format_text(ws)
    ws.cell(row=ws.max_row + 1, column=1).value = "Event Time"
    ws.cell(row=ws.max_row, column=2).value = f"{event_time}"
    ws.cell(row=ws.max_row + 1, column=1).value = "Time Until"
    ws.cell(row=ws.max_row, column=2).value = f"{time_until}"

    ws.cell(row=ws.max_row + 2, column=1).value = "Money Line"
    format_text(ws)
    for data in moneyline:
        ws.append(data)

    ws.cell(row=ws.max_row + 2, column=1).value = "Handicap"
    format_text(ws)
    for data in handicap:
        ws.append(data)

    ws.cell(row=ws.max_row + 2, column=1).value = "Total"
    ws.cell(row=ws.max_row, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=3)
    for data in total:
        ws.append(data)

    wb.save(filename='ODDS_OUTPUT.xlsx')


def is_today(tag):
    if "data-test-id" in tag.attrs and tag['data-test-id'] == "Events.DateBar":
        return tag.find("span", attrs={"class": "bold"}).find_next("span").string == "Today"


def get_all_odds():
    global soup
    soup = update_soup()

    if soup.find(is_today) is None:
        return

    rows = soup.find_all("div", attrs={"data-test-id": ["Events.DateBar", "Event.Row"]})
    date_rows = soup.find_all("div", attrs={"data-test-id": "Events.DateBar"})
    dates_today_index = date_rows.index(soup.find(is_today))
    today_index = rows.index(soup.find(is_today))
    if dates_today_index == len(date_rows) - 1:
        tomorrow_index = len(rows)
    else:
        tomorrow_index = rows.index(date_rows[dates_today_index + 1])

    for i in range(today_index + 1, tomorrow_index, 1):
        if len(rows[i].find_all("span", attrs={"class": "price"})) == 7:
            event_datetime = str(rows[i].find("script")).split(',')[3].split('"')[3]
            event_datetime_utc = dateutil.parser.parse(event_datetime)
            times_list.append(event_datetime_utc)


def get_event_odds(time):
    rows = soup.find_all("div", attrs={"data-test-id": ["Events.DateBar", "Event.Row"]})
    date_rows = soup.find_all("div", attrs={"data-test-id": "Events.DateBar"})
    dates_today_index = date_rows.index(soup.find(is_today))
    today_index = rows.index(soup.find(is_today))
    if dates_today_index == len(date_rows) - 1:
        tomorrow_index = len(rows)
    else:
        tomorrow_index = rows.index(date_rows[dates_today_index + 1])

    start_temp_driver()

    tm.sleep(14 * (5 - times_list.count(time)))

    for i in range(today_index + 1, tomorrow_index, 1):
        event_datetime = str(rows[i].find("script")).split(',')[3].split('"')[3]
        event_datetime_utc = dateutil.parser.parse(event_datetime)

        if event_datetime_utc == time:
            event_url = str(rows[i].find("script")).split(',')[4].split('"')[3]
            teams = rows[i].find_all("span", attrs={"class": re.compile("vRjBw")})

            event_soup = pull_event_odds(event_url)
            event_rows = event_soup.find_all("div", attrs={"data-test-id": "Collapse"}, limit=3)

            time_until_event = event_datetime_utc - datetime.datetime.now(tzutc())

            if len(event_rows) == 3:
                moneyline_prices, handicap_prices, handicap_points, total_prices, total_points = [], [[], []], [[], []], [[], []], [[], []]

                for price in event_rows[0].find_all("span", class_="price"):
                    moneyline_prices.append(float(price.string))
                for i in range(0, len(event_rows[1].find_all("span", class_="price")), 2):
                    handicap_prices[0].append(float(event_rows[1].find_all("span", class_="price")[i].string))
                    handicap_prices[1].append(float(event_rows[1].find_all("span", class_="price")[i + 1].string))
                for i in range(0, len(event_rows[1].find_all("span", class_="label")), 2):
                    handicap_points[0].append(float(event_rows[1].find_all("span", class_="label")[i].string))
                    handicap_points[1].append(float(event_rows[1].find_all("span", class_="label")[i + 1].string))
                for i in range(0, len(event_rows[2].find_all("span", class_="price")), 2):
                    total_prices[0].append(float(event_rows[2].find_all("span", class_="price")[i].string))
                    total_prices[1].append(float(event_rows[2].find_all("span", class_="price")[i + 1].string))
                for i in range(0, len(event_rows[2].find_all("span", class_="label")), 2):
                    total_points[0].append(float(event_rows[2].find_all("span", class_="label")[i].string[5:]))
                    total_points[1].append(float(event_rows[2].find_all("span", class_="label")[i + 1].string[6:]))

                moneyline = [[teams[0].string, "DRAW", teams[1].string], moneyline_prices]
                handicap = [[teams[0].string], handicap_points[0], handicap_prices[0], [teams[1].string],
                            handicap_points[1], handicap_prices[1]]
                total = [["OVER"], total_points[0], total_prices[0], ["UNDER"], total_points[1], total_prices[1]]

                if len(moneyline_prices) > 0 and len(handicap_prices) > 0 and len(total_prices) > 0:
                    print_odds(moneyline, handicap, total, event_datetime_utc, time_until_event)

    close_temp_driver()


def main():
    get_all_odds()
    print("Games Times for Today:")
    for time in times_list:
        print(time)

    time_delta = datetime.timedelta(seconds=83)

    while True:
        if len(times_list) == 0:
            break
        for time in times_list:
            if time - datetime.datetime.now(tzutc()) <= time_delta:
                get_event_odds(time)
                times_list[:] = [x for x in times_list if x != time]


if __name__ == '__main__':
    main()
